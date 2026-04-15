"""Metadata extraction service — uses Gemini structured output to extract
document metadata based on the admin-configured schema."""

import json
import logging
import time

from google import genai
from google.genai import types

from app.services.settings import get_llm_api_key, get_llm_model, get_metadata_schema

logger = logging.getLogger(__name__)

_client_cache = {"key": None, "client": None}


def _get_client() -> genai.Client:
    key = get_llm_api_key()
    if _client_cache["key"] != key:
        _client_cache["key"] = key
        _client_cache["client"] = genai.Client(api_key=key)
    return _client_cache["client"]


def build_extraction_prompt(text: str, schema: list[dict]) -> str:
    """Build the extraction prompt dynamically from the schema."""
    fields_desc = []
    for field in schema:
        req = "required" if field.get("required") else "optional"
        fields_desc.append(f"- {field['name']} ({field['type']}, {req}): {field.get('description', '')}")

    fields_str = "\n".join(fields_desc)

    # Truncate text to ~4000 words
    words = text.split()
    if len(words) > 4000:
        truncated = " ".join(words[:4000])
    else:
        truncated = text

    return f"""Analyze the following document and extract metadata for each field.

Fields to extract:
{fields_str}

Return a JSON object with these exact keys. For required fields, always provide a value.
For optional fields, use null if not applicable.
For "list" type, return an array of strings.
For "boolean" type, return true or false.
For "number" type, return a numeric value or null.
For "date" type, return an ISO date string (YYYY-MM-DD) or null.

Document text:
{truncated}"""


def build_response_schema(schema: list[dict]) -> dict:
    """Convert field definitions into a JSON Schema dict for Gemini's response_schema."""
    properties = {}
    required = []

    for field in schema:
        field_type = field.get("type", "text")
        if field_type == "text":
            properties[field["name"]] = {"type": "STRING"}
        elif field_type == "list":
            properties[field["name"]] = {"type": "ARRAY", "items": {"type": "STRING"}}
        elif field_type == "boolean":
            properties[field["name"]] = {"type": "BOOLEAN"}
        elif field_type == "number":
            properties[field["name"]] = {"type": "NUMBER"}
        elif field_type == "date":
            properties[field["name"]] = {"type": "STRING"}
        else:
            properties[field["name"]] = {"type": "STRING"}

        if field.get("required"):
            required.append(field["name"])

    return {
        "type": "OBJECT",
        "properties": properties,
        "required": required,
    }


def _get_fallback(schema: list[dict]) -> dict:
    """Return a fallback metadata dict with defaults for all fields."""
    result = {}
    for field in schema:
        field_type = field.get("type", "text")
        if field_type == "text":
            result[field["name"]] = "unknown" if field.get("required") else None
        elif field_type == "list":
            result[field["name"]] = []
        elif field_type == "boolean":
            result[field["name"]] = False
        elif field_type == "number":
            result[field["name"]] = None
        elif field_type == "date":
            result[field["name"]] = None
        else:
            result[field["name"]] = None
    return result


def enrich_tabular_text(text: str, file_name: str, file_content: bytes = None) -> str:
    """For CSV/XLSX files, prepend filename and column headers so the LLM has
    enough context to extract meaningful metadata instead of 'unknown'."""
    import csv, io, os

    ext = os.path.splitext(file_name)[1].lower() if file_name else ""
    if ext not in (".csv", ".xlsx", ".xlsm") or not file_content:
        return text

    parts = [f"File name: {file_name}\n"]

    try:
        if ext == ".csv":
            decoded = file_content.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(decoded))
            headers = reader.fieldnames or []
            parts.append(f"Columns: {', '.join(headers)}\n")
            sample_rows = []
            for i, row in enumerate(reader):
                if i >= 5:
                    break
                sample_rows.append(", ".join(f"{k}={v}" for k, v in row.items()))
            if sample_rows:
                parts.append("Sample rows:\n" + "\n".join(sample_rows) + "\n")

        elif ext in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            for sheet in wb.worksheets:
                rows = list(sheet.iter_rows(max_row=6, values_only=True))
                if not rows:
                    continue
                headers = [str(h) if h else "" for h in rows[0]]
                parts.append(f"Sheet '{sheet.title}' columns: {', '.join(headers)}\n")
                for row in rows[1:]:
                    parts.append(", ".join(f"{h}={v}" for h, v in zip(headers, row) if v is not None))
                parts.append("")
            wb.close()
    except Exception:
        pass  # Fall through to raw text

    parts.append(text)
    return "\n".join(parts)


def extract_metadata(text: str, file_name: str = "", file_content: bytes = None) -> dict:
    """Extract structured metadata from document text using Gemini."""
    text = enrich_tabular_text(text, file_name, file_content)
    schema = get_metadata_schema()
    prompt = build_extraction_prompt(text, schema)
    response_schema = build_response_schema(schema)

    client = _get_client()
    model = get_llm_model()

    for attempt in range(5):
        try:
            response = client.models.generate_content(
                model=model,
                contents=prompt,
                config=types.GenerateContentConfig(
                    response_mime_type="application/json",
                    response_schema=response_schema,
                ),
            )
            result = json.loads(response.text)
            return result
        except Exception as e:
            if attempt == 4:
                logger.error(f"Metadata extraction failed after 5 attempts: {e}")
                return _get_fallback(schema)
            err_str = str(e)
            if "429" in err_str or "RESOURCE_EXHAUSTED" in err_str:
                wait = 5 * (2 ** attempt)  # 5, 10, 20, 40, 80s
            else:
                wait = 2 ** attempt
            logger.warning(f"Metadata extraction attempt {attempt + 1} failed: {e}. Retrying in {wait}s...")
            time.sleep(wait)

    return _get_fallback(schema)


def extract_query_filters(query: str) -> dict | None:
    """Use LLM to infer metadata filters from a user's natural language query.
    Returns a dict of filters for retrieve_chunks(), or None if no filters inferred."""
    schema = get_metadata_schema()
    if not schema:
        return None

    fields_desc = []
    for field in schema:
        fields_desc.append(f"- {field['name']} ({field['type']}): {field.get('description', '')}")
    fields_str = "\n".join(fields_desc)

    prompt = f"""Given a user's search query, extract metadata filters that would help narrow document retrieval.
Only extract filters you are highly confident about based on the query. If no filters can be inferred, return an empty JSON object {{}}.

Available metadata fields:
{fields_str}

Rules:
- Only include fields where the query clearly implies a filter value.
- For "text" fields, return the most likely value as a string.
- For "list" fields like "keywords" or "entities", return an array with 1-3 matching terms.
- For "boolean" fields, return true or false.
- For "date" fields, return YYYY-MM-DD format if a specific date is mentioned. For year-only mentions like "in 2023", do NOT include a date filter.
- For "number" fields, return a numeric value.
- Omit fields where you cannot confidently determine a value.
- Prefer keywords and entities filters for topical queries.
- Do NOT filter on document_type, language, or summary unless the user explicitly mentions them.

User query: {query}"""

    client = _get_client()
    model = get_llm_model()

    try:
        response = client.models.generate_content(
            model=model,
            contents=prompt,
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
            ),
        )
        result = json.loads(response.text)

        # Validate keys — discard any not in schema
        valid_names = {f["name"] for f in schema}
        filtered = {k: v for k, v in result.items() if k in valid_names and v is not None}

        # Discard empty values
        cleaned = {}
        for k, v in filtered.items():
            if isinstance(v, list) and len(v) == 0:
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            cleaned[k] = v

        return cleaned if cleaned else None
    except Exception as e:
        logger.warning(f"Query filter extraction failed (non-fatal): {e}")
        return None
